import json
import os
import re

from bs4 import BeautifulSoup
import pandas as pd
import charset_normalizer
import docx2txt
import numpy as np
import openpyxl
import pandas as pd
import pdfplumber
import requests
import torch
import xlrd
from PIL import Image, ImageOps, ImageSequence
from markdownify import markdownify as md
from ..config import current_dir_path

file_path = os.path.join(current_dir_path, "file")
programming_languages_extensions = [".py", ".js", ".java", ".c", ".cpp", ".html", ".css", ".sql", ".r", ".swift"]


def read_one(path):
    text = ""
    if path.endswith(".docx") or path.endswith(".DOCX"):
        text += docx2txt.process(path)
    elif path.endswith(".md") or path.endswith(".MD"):
        with open(path, "r", encoding="utf-8") as f:
            text += f.read()
    elif path.endswith(".pdf") or path.endswith(".PDF"):
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                text += page.extract_text()
    elif path.endswith(".xlsx") or path.endswith(".XLSX"):
        workbook = openpyxl.load_workbook(path)
        for sheet in workbook.worksheets:
            # 检查工作表是否至少有一行数据
            if sheet.max_row > 1:  # 至少有一行数据（不包括表头）
                # 获取工作表名称
                text += f"## {sheet.title} 的内容\n"

                # 获取表头
                headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if headers:
                    text += "|" + " | ".join([" " if cell is None else str(cell) for cell in headers]) + "|\n"
                    text += "|" + " | ".join(["---"] * len(headers)) + "|\n"
                else:
                    text += "没有找到表头。\n"
                    continue

                # 获取数据行
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if any(cell is not None for cell in row):
                        text += "|" + " | ".join([" " if cell is None else str(cell) for cell in row]) + "|\n"
                    else:
                        # 如果整行都是空的，则停止读取当前工作表
                        break
    elif path.endswith(".xls") or path.endswith(".XLS"):
        workbook = xlrd.open_workbook(path)
        for sheet_index in range(workbook.nsheets):
            sheet = workbook.sheet_by_index(sheet_index)
            # 检查工作表是否为空
            if sheet.nrows > 0:
                text += f"## {sheet.name} 的内容\n"
                text += "| " + " | ".join(sheet.row_values(0)) + " |\n"  # 添加表头
                text += "| " + " | ".join(["---"] * sheet.ncols) + " |\n"  # 添加分隔符
                for row_num in range(1, sheet.nrows):
                    text += "| " + " | ".join([str(cell) for cell in sheet.row_values(row_num)]) + " |\n"
    elif path.endswith(".csv") or path.endswith(".CSV"):
        # 检测文件编码
        with open(path, "rb") as file:
            content = file.read()
        result = charset_normalizer.from_bytes(content)
        encoding = result.best().encoding if result else 'utf-8'
        # 读取 CSV 文件
        df = pd.read_csv(path, encoding=encoding)
        text += df.to_markdown(index=True)
    elif path.endswith(".txt") or path.endswith(".TXT"):
        with open(path, "r", encoding="utf-8") as f:
            text += f.read()
    elif path.endswith(".json") or path.endswith(".JSON"):
        with open(path, "r", encoding="utf-8") as f:
            text += json.dumps(json.load(f), ensure_ascii=False, indent=4)
    elif any(path.endswith(extension) for extension in programming_languages_extensions):
        try:
            with open(path, "r", encoding="utf-8") as file:
                text += file.read()
        except UnicodeDecodeError:
            with open(path, "r", encoding="latin-1") as file:
                text += file.read()
    out = [{"source": path,"paragraph_index":"full text" , "file_content": text}]
    return out


file_path = os.path.join(current_dir_path, "file")


class load_file:
    @classmethod
    def INPUT_TYPES(s):
        paths = [f for f in os.listdir(file_path)]
        return {
            "required": {
                "absolute_path": ("STRING", {"default": ""}),
                "relative_path": (paths, {"default": "test.txt"}),
                "is_enable": ("BOOLEAN", {"default": True}),
            },
            "optional": {},
        }

    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("file_content",)

    FUNCTION = "file"

    # OUTPUT_NODE = False

    CATEGORY = "大模型派对（llm_party）/知识库（knowbase）"

    def file(self, relative_path, absolute_path="", is_enable=True):
        if is_enable == False:
            return (None,)
        if absolute_path != "":
            path = absolute_path
        else:
            path = os.path.join(file_path, relative_path)
        out = read_one(path)
        out = json.dumps(out, ensure_ascii=False, indent=4)
        return (out,)


class load_file_folder:
    @classmethod
    def INPUT_TYPES(s):
        return {
            "required": {
                "folder_path": ("STRING", {"default": "C://Users/"}),
                "is_enable": ("BOOLEAN", {"default": True}),
            },
            "optional": {
                "related_characters": ("STRING", {"default": ""}),
            },
        }

    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("file_content",)

    FUNCTION = "file"

    # OUTPUT_NODE = False

    CATEGORY = "大模型派对（llm_party）/知识库（knowbase）"

    def file(self, folder_path, related_characters="", is_enable=True):
        if is_enable == False:
            return (None,)
        # 获取文件夹中的所有文件，并读取到字符串中
        out = []
        for root, dirs, files in os.walk(folder_path):
            if related_characters != "":
                for file in files:
                    # 如果文件名包含相关字符
                    if related_characters in file:
                        file_path = os.path.join(root, file)
                        out.extend(read_one(file_path))
            else:
                for file in files:
                    file_path = os.path.join(root, file)
                    out.extend(read_one(file_path))
        out = json.dumps(out, ensure_ascii=False, indent=4)
        return (out,)


class load_url:
    @classmethod
    def INPUT_TYPES(s):
        return {
            "required": {
                "url": ("STRING", {"default": "https://example.com"}),
                "is_enable": ("BOOLEAN", {"default": True}),
            },
            "optional": {
                "with_jina": ("BOOLEAN", {"default": True}),
            },
        }

    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("file_content",)

    FUNCTION = "file"

    # OUTPUT_NODE = False

    CATEGORY = "大模型派对（llm_party）/知识库（knowbase）"

    def file(self, url, with_jina=True, is_enable=True):
        if not is_enable:
            return (None,)
        try:
            jina = "https://r.jina.ai/"
            if with_jina:
                url = jina + url

            response = requests.get(url, timeout=10)
            response.raise_for_status()  # 确保请求成功

            # 使用 charset_normalizer 检测编码
            detected_encoding = charset_normalizer.detect(response.content)['encoding']
            response.encoding = detected_encoding if detected_encoding else 'utf-8'
        except requests.exceptions.RequestException as e:
            print(f"请求发生错误: {e}")
            return (None,)
        
        out = response.text
        print(out)  # Debugging: Check the HTML content
        
        if not with_jina:
            # 使用 BeautifulSoup 解析 HTML
            soup = BeautifulSoup(out, 'html.parser')
            # 提取主要内容
            main_content = soup.get_text()
            # 将 HTML 转换为 Markdown
            out = md(main_content, convert=['p', 'h1', 'h2', 'h3', 'a', 'img'], heading_style="ATX", bullets="*+-", strong_em_symbol="ASTERISK")
            # 去掉多余的换行符
            out = re.sub(r'\n+', '\n', out)
        out = [{"source": url,"paragraph_index":"full text" , "file_content": out}]
        out = json.dumps(out, ensure_ascii=False, indent=4)
        return (out,)


# 定义一个函数，将图片路径转换为与 save_images 函数兼容的格式
def convert_image_to_compatible_format(image_path):
    # 从给定路径加载图片
    img = Image.open(image_path).convert("RGB")

    # 将图片转换为numpy数组并归一化
    img_array = np.asarray(img) / 255.0

    # 将numpy数组转换为torch张量
    img_tensor = torch.tensor(img_array).permute(2, 0, 1).unsqueeze(0)  # 调整维度顺序并添加批次维度

    # 返回预期格式的张量
    return img_tensor


class start_workflow:
    @classmethod
    def INPUT_TYPES(s):
        return {
            "required": {},
            "optional": {
                "file_content": ("STRING", {"forceInput": True}),
                "image_input1": ("IMAGE", {}),
                "image_input2": ("IMAGE", {}),
                "file_path": ("STRING", {"default": ""}),
                "img_path1": ("STRING", {"default": "", "image_upload": True}),
                "img_path2": ("STRING", {"default": "", "image_upload": True}),
                "system_prompt": ("STRING", {"default": "你是一个强大的智能助手"}),
                "user_prompt": ("STRING", {"default": "你好"}),
                "positive_prompt": ("STRING", {"default": ""}),
                "negative_prompt": ("STRING", {"default": ""}),
                "model_name": ("STRING", {"default": ""}),
                "user_history": ("STRING", {"default": ""}),
            },
        }

    RETURN_TYPES = (
        "STRING",
        "IMAGE",
        "IMAGE",
        "STRING",
        "STRING",
        "STRING",
        "STRING",
        "STRING",
        "STRING",
    )
    RETURN_NAMES = (
        "file_content",
        "image1",
        "image2",
        "system_prompt",
        "user_prompt",
        "positive_prompt",
        "negative_prompt",
        "model_name",
        "user_history",
    )

    FUNCTION = "load_all"

    # OUTPUT_NODE = False

    CATEGORY = "大模型派对（llm_party）/工作流（workflow）"

    def load_all(
        self,
        file_content="",
        image_input1="",
        image_input2="",
        file_path="",
        img_path1="",
        img_path2="",
        system_prompt="你是一个强大的智能助手",
        user_prompt="你好",
        positive_prompt="",
        negative_prompt="",
        model_name="",
        user_history="",
    ):
        file_out = []
        if file_content is not None and file_content != "":
            file_out += file_content
        if file_path is not None and file_path != "":
            if os.path.isdir(file_path):
                for path in os.listdir(file_path):
                    path = os.path.join(file_path, path)
                    file_out.extend(read_one(path))
            else:
                print(file_path)
                file_out = read_one(file_path)
                print(file_out)
        file_out = json.dumps(file_out, ensure_ascii=False, indent=4)
        img_out = []
        if image_input1 is not None:
            img_out=image_input1
        img_out2 = []
        if image_input2 is not None:
            img_out2=image_input2
        if img_path1 is not None and img_path1 != "":
            img_out = []
            # 检查img_path是否是一个目录
            if os.path.isdir(img_path1):
                # 遍历目录中的所有文件
                for filename in os.listdir(img_path1):
                    # 检查文件是否是图片
                    if filename.lower().endswith((".png", ".jpg", ".jpeg", ".tiff", ".bmp", ".gif")):
                        img_full_path = os.path.join(img_path1, filename)
                        img = Image.open(img_full_path)
                        img = ImageOps.exif_transpose(img)
                        if img.mode == "I":
                            img = img.point(lambda i: i * (1 / 256)).convert("L")
                        image = img.convert("RGB")
                        image = np.array(image).astype(np.float32) / 255.0
                        image = torch.from_numpy(image).unsqueeze(0)
                        img_out.append(image)
            else:
                img = Image.open(img_path1)
                for i in ImageSequence.Iterator(img):
                    i = ImageOps.exif_transpose(i)
                    if i.mode == "I":
                        i = i.point(lambda i: i * (1 / 256)).convert("L")
                    image = i.convert("RGB")
                    image = np.array(image).astype(np.float32) / 255.0
                    image = torch.from_numpy(image).unsqueeze(0)
                    img_out.append(image)

            if len(img_out) > 1:
                img_out = torch.cat(img_out, dim=0)
            elif img_out:
                img_out = img_out[0]

        if img_path2 is not None and img_path2 != "":
            img_out2 = []
            # 检查img_path是否是一个目录
            if os.path.isdir(img_path2):
                # 遍历目录中的所有文件
                for filename in os.listdir(img_path2):
                    # 检查文件是否是图片
                    if filename.lower().endswith((".png", ".jpg", ".jpeg", ".tiff", ".bmp", ".gif")):
                        img_full_path = os.path.join(img_path2, filename)
                        img = Image.open(img_full_path)
                        img = ImageOps.exif_transpose(img)
                        if img.mode == "I":
                            img = img.point(lambda i: i * (1 / 256)).convert("L")
                        image = img.convert("RGB")
                        image = np.array(image).astype(np.float32) / 255.0
                        image = torch.from_numpy(image).unsqueeze(0)
                        img_out.append(image)
            else:
                img = Image.open(img_path2)
                for i in ImageSequence.Iterator(img):
                    i = ImageOps.exif_transpose(i)
                    if i.mode == "I":
                        i = i.point(lambda i: i * (1 / 256)).convert("L")
                    image = i.convert("RGB")
                    image = np.array(image).astype(np.float32) / 255.0
                    image = torch.from_numpy(image).unsqueeze(0)
                    img_out2.append(image)

            if len(img_out2) > 1:
                img_out2 = torch.cat(img_out2, dim=0)
            elif img_out2:
                img_out2 = img_out2[0]

        system_out = system_prompt
        user_out = user_prompt
        positive_out = positive_prompt
        negative_out = negative_prompt
        model_name_out = model_name
        user_history_out = user_history
        return (
            file_out,
            img_out,
            img_out2,
            system_out,
            user_out,
            positive_out,
            negative_out,
            model_name_out,
            user_history_out,
        )


class load_img_path:
    @classmethod
    def INPUT_TYPES(s):
        return {
            "required": {
                "img_path": ("STRING", {"default": "", "image_upload": True}),
            },
        }

    RETURN_TYPES = ("IMAGE",)
    RETURN_NAMES = ("image",)

    FUNCTION = "load_all"

    # OUTPUT_NODE = False

    CATEGORY = "大模型派对（llm_party）/图片（image）"

    def load_all(self, img_path):
        img_out = []
        if img_path is not None and img_path != "":
            # 检查img_path是否是一个目录
            if os.path.isdir(img_path):
                # 遍历目录中的所有文件
                for filename in os.listdir(img_path):
                    # 检查文件是否是图片
                    if filename.lower().endswith((".png", ".jpg", ".jpeg", ".tiff", ".bmp", ".gif")):
                        img_full_path = os.path.join(img_path, filename)
                        img = Image.open(img_full_path)
                        img = ImageOps.exif_transpose(img)
                        if img.mode == "I":
                            img = img.point(lambda i: i * (1 / 256)).convert("L")
                        image = img.convert("RGB")
                        image = np.array(image).astype(np.float32) / 255.0
                        image = torch.from_numpy(image).unsqueeze(0)
                        img_out.append(image)
            else:
                img = Image.open(img_path)
                for i in ImageSequence.Iterator(img):
                    i = ImageOps.exif_transpose(i)
                    if i.mode == "I":
                        i = i.point(lambda i: i * (1 / 256)).convert("L")
                    image = i.convert("RGB")
                    image = np.array(image).astype(np.float32) / 255.0
                    image = torch.from_numpy(image).unsqueeze(0)
                    img_out.append(image)

        if len(img_out) > 1:
            img_out = torch.cat(img_out, dim=0)
        elif img_out:
            img_out = img_out[0]
        if img_out == []:
            img_out = None
        return (img_out,)
