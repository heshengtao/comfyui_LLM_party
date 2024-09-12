import datetime
import hashlib
import json
import locale
import shutil
import pandas as pd
import charset_normalizer
import docx2txt
import numpy as np
import openpyxl
import pandas as pd
import pdfplumber
import xlrd
import os
import json

current_dir_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
file_path = os.path.join(current_dir_path, "file")
programming_languages_extensions = [".py", ".js", ".java", ".c", ".cpp", ".html", ".css", ".sql", ".r", ".swift"]


def read_one(path):
    text = ""
    if path.endswith(".docx"):
        text += docx2txt.process(path)
    elif path.endswith(".md"):
        with open(path, "r", encoding="utf-8") as f:
            text += f.read()
    elif path.endswith(".pdf"):
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                text += page.extract_text()
    elif path.endswith(".xlsx"):
        workbook = openpyxl.load_workbook(path)
        for sheet in workbook.worksheets:
            # 检查工作表是否至少有一行数据
            if sheet.max_row > 1:  # 至少有一行数据（不包括表头）
                # 获取工作表名称
                text += f"## {sheet.title} 的内容\n"

                # 获取表头
                headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if headers:
                    text += "|" + " | ".join([" " if cell is None else str(cell) for cell in headers]) + "|\n"
                    text += "|" + " | ".join(["---"] * len(headers)) + "|\n"
                else:
                    text += "没有找到表头。\n"
                    continue

                # 获取数据行
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if any(cell is not None for cell in row):
                        text += "|" + " | ".join([" " if cell is None else str(cell) for cell in row]) + "|\n"
                    else:
                        # 如果整行都是空的，则停止读取当前工作表
                        break
    elif path.endswith(".xls"):
        workbook = xlrd.open_workbook(path)
        for sheet_index in range(workbook.nsheets):
            sheet = workbook.sheet_by_index(sheet_index)
            # 检查工作表是否为空
            if sheet.nrows > 0:
                text += f"## {sheet.name} 的内容\n"
                text += "| " + " | ".join(sheet.row_values(0)) + " |\n"  # 添加表头
                text += "| " + " | ".join(["---"] * sheet.ncols) + " |\n"  # 添加分隔符
                for row_num in range(1, sheet.nrows):
                    text += "| " + " | ".join([str(cell) for cell in sheet.row_values(row_num)]) + " |\n"
    elif path.endswith(".csv"):
        # 检测文件编码
        with open(path, "rb") as file:
            content = file.read()
        result = charset_normalizer.from_bytes(content)
        encoding = result.best().encoding if result else 'utf-8'
        # 读取 CSV 文件
        df = pd.read_csv(path, encoding=encoding)
        text += df.to_markdown(index=True)
    elif path.endswith(".txt"):
        with open(path, "r", encoding="utf-8") as f:
            text += f.read()
    elif path.endswith(".json"):
        with open(path, "r", encoding="utf-8") as f:
            text += json.dumps(json.load(f), ensure_ascii=False, indent=4)
    elif any(path.endswith(extension) for extension in programming_languages_extensions):
        try:
            with open(path, "r", encoding="utf-8") as file:
                text += file.read()
        except UnicodeDecodeError:
            with open(path, "r", encoding="latin-1") as file:
                text += file.read()

    return text

global_file_path = ""

def file_control(file_path, mode="w", text_or_path=""):
    global global_file_path
    # 规范化路径
    file_path = os.path.abspath(file_path)
    global_file_path = os.path.abspath(global_file_path)
    
    if not file_path.startswith(global_file_path):
        return "文件路径不在指定目录下"
    
    if mode in ["w", "a"]:
        try:
            with open(file_path, mode, encoding="utf-8") as f:
                f.write(text_or_path + "\n")
        except Exception as e:
            raise ValueError(f"写入文件失败: {e}")
    elif mode == "r":
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                res = f.read()
            return "读取成功,内容为：" + res
        except Exception as e:
            raise ValueError(f"读取文件失败: {e}")
    elif mode == "d":
        try:
            os.remove(file_path)
        except Exception as e:
            raise ValueError(f"删除文件失败: {e}")
    elif mode == "m":
        text_or_path = os.path.abspath(text_or_path)
        if not text_or_path.startswith(global_file_path):
            return "移动后的文件路径不在指定目录下"
        try:
            shutil.move(file_path, text_or_path)
        except Exception as e:
            raise ValueError(f"移动文件失败: {e}")
    elif mode == "c":
        text_or_path = os.path.abspath(text_or_path)
        if not text_or_path.startswith(global_file_path):
            return "复制后的文件路径不在指定目录下"
        try:
            shutil.copy(file_path, text_or_path)
        except Exception as e:
            raise ValueError(f"复制文件失败: {e}")
    elif mode == "n":
        text_or_path = os.path.abspath(text_or_path)
        if not text_or_path.startswith(global_file_path):
            return "重命名后的文件路径不在指定目录下"
        try:
            os.rename(file_path, text_or_path)
        except Exception as e:
            raise ValueError(f"重命名文件失败: {e}")
    return "操作成功"


def get_file_tree(folder_path):
    def build_tree(path):
        tree = {}
        for item in os.listdir(path):
            item_path = os.path.join(path, item)
            if os.path.isdir(item_path):
                tree[item] = build_tree(item_path)
            else:
                tree[item] = item
        return tree

    return {folder_path: build_tree(folder_path)}


class files_control_tool:
    @classmethod
    def INPUT_TYPES(s):
        return {
            "required": {
                "folder_path": ("STRING", {"default": "folder path/to/you want to llm control"}),
                "is_enable": ("BOOLEAN", {"default": True}),
            }
        }

    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("tool",)

    FUNCTION = "time"

    # OUTPUT_NODE = False

    CATEGORY = "大模型派对（llm_party）/工具（tools）"

    def time(self, folder_path, is_enable=True):
        if is_enable == False:
            return (None,)
        print("Please don't let LLM control your important documents!")
        global global_file_path
        global_file_path=folder_path
        if global_file_path=="" or global_file_path==None:
            print("folder path must be set！")
            return (None,)
        file_tree = get_file_tree(folder_path)
        file_tree = json.dumps(file_tree, ensure_ascii=False, indent=4)
        output = [
            {
                "type": "function",
                "function": {
                    "name": "file_control",
                    "description": f"用于读取、写入、追加、删除、移动、重命名、复制在{folder_path}下的文件，注意！你不能操作其他路径下的文件，写入、追加时，最好是对一个新文件执行，除非用户给定了文件。文件树如下：\n{file_tree}",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "file_path": {
                                "type": "string",
                                "description": "在文件的绝对路径",
                            },
                            "mode": {
                                "type": "string",
                                "description": "w:写入，a:追加，r:读取，d:删除，m:移动，n:重命名，c:复制",
                            },
                            "text_or_path": {
                                "type": "string",
                                "description": "写入或追加时，text_or_path为要写入或追加的文本,读取和删除时不需要,移动时，text_or_path为移动到的绝对路径，重命名时，text_or_path为新的文件名的绝对路径，复制时，text_or_path为复制到的绝对路径",
                            }
                        },
                        "required": ["file_path"],
                    },
                },
            }
        ]
        out = json.dumps(output, ensure_ascii=False)
        return (out,)
    @classmethod
    def IS_CHANGED(s):
        # 生成当前时间的哈希值
        hash_value = hashlib.md5(str(datetime.datetime.now()).encode()).hexdigest()
        return hash_value

_TOOL_HOOKS = ["file_control"]
NODE_CLASS_MAPPINGS = {"files_control_tool": files_control_tool}
# 获取系统语言
lang = locale.getdefaultlocale()[0]
import os
import sys
current_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
config_path = os.path.join(current_dir, "config.ini")
import configparser
config = configparser.ConfigParser()
config.read(config_path)
try:
    language = config.get("API_KEYS", "language")
except:
    language = ""
if language == "zh_CN" or language=="en_US":
    lang=language
if lang == "zh_CN":
    NODE_DISPLAY_NAME_MAPPINGS = {"files_control_tool": "(危险！)本地文件控制工具"}
else:
    NODE_DISPLAY_NAME_MAPPINGS = {"files_control_tool": "(Danger!) local file control tool"}

if __name__ == "__main__":
    # 示例用法
    folder_path = 'D:\AI\AIhuitu\Blender_ComfyUI\Blender_ComfyUI_Mini\ComfyUI\custom_nodes\comfyui_LLM_party\KG'
    file_tree = get_file_tree(folder_path)

    # 将文件树转换为 JSON 格式
    file_tree_json = json.dumps(file_tree, indent=4, ensure_ascii=False)
    print(file_tree_json)
