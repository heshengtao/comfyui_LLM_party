import datetime
import hashlib
import json
import locale
import shutil
import pandas as pd
import charset_normalizer
import docx2txt
import numpy as np
import openpyxl
import pandas as pd
import pdfplumber
import xlrd
import os
import json

current_dir_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
file_path = os.path.join(current_dir_path, "file")
programming_languages_extensions = [".py", ".js", ".java", ".c", ".cpp", ".html", ".css", ".sql", ".r", ".swift"]


def read_one(path):
    text = ""
    if path.endswith(".docx"):
        text += docx2txt.process(path)
    elif path.endswith(".md"):
        with open(path, "r", encoding="utf-8") as f:
            text += f.read()
    elif path.endswith(".pdf"):
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                text += page.extract_text()
    elif path.endswith(".xlsx"):
        workbook = openpyxl.load_workbook(path)
        for sheet in workbook.worksheets:
            # 检查工作表是否至少有一行数据
            if sheet.max_row > 1:  # 至少有一行数据（不包括表头）
                # 获取工作表名称
                text += f"## {sheet.title} 的内容\n"

                # 获取表头
                headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if headers:
                    text += "|" + " | ".join([" " if cell is None else str(cell) for cell in headers]) + "|\n"
                    text += "|" + " | ".join(["---"] * len(headers)) + "|\n"
                else:
                    text += "没有找到表头。\n"
                    continue

                # 获取数据行
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if any(cell is not None for cell in row):
                        text += "|" + " | ".join([" " if cell is None else str(cell) for cell in row]) + "|\n"
                    else:
                        # 如果整行都是空的，则停止读取当前工作表
                        break
    elif path.endswith(".xls"):
        workbook = xlrd.open_workbook(path)
        for sheet_index in range(workbook.nsheets):
            sheet = workbook.sheet_by_index(sheet_index)
            # 检查工作表是否为空
            if sheet.nrows > 0:
                text += f"## {sheet.name} 的内容\n"
                text += "| " + " | ".join(sheet.row_values(0)) + " |\n"  # 添加表头
                text += "| " + " | ".join(["---"] * sheet.ncols) + " |\n"  # 添加分隔符
                for row_num in range(1, sheet.nrows):
                    text += "| " + " | ".join([str(cell) for cell in sheet.row_values(row_num)]) + " |\n"
    elif path.endswith(".csv"):
        # 检测文件编码
        with open(path, "rb") as file:
            content = file.read()
        result = charset_normalizer.from_bytes(content)
        encoding = result.best().encoding if result else 'utf-8'
        # 读取 CSV 文件
        df = pd.read_csv(path, encoding=encoding)
        text += df.to_markdown(index=True)
    elif path.endswith(".txt"):
        with open(path, "r", encoding="utf-8") as f:
            text += f.read()
    elif path.endswith(".json"):
        with open(path, "r", encoding="utf-8") as f:
            text += json.dumps(json.load(f), ensure_ascii=False, indent=4)
    elif any(path.endswith(extension) for extension in programming_languages_extensions):
        try:
            with open(path, "r", encoding="utf-8") as file:
                text += file.read()
        except UnicodeDecodeError:
            with open(path, "r", encoding="latin-1") as file:
                text += file.read()
    out = [{"source": path,"paragraph_index":"full text" , "file_content": text}]
    return out

global_file_path = ""

def file_read(file_path):
    global global_file_path
    # 规范化路径
    file_path = os.path.abspath(file_path)
    global_file_path = os.path.abspath(global_file_path)
    
    if not file_path.startswith(global_file_path):
        return "文件路径不在指定目录下"
    
    def ensure_directory_exists(path):
        folder_path = os.path.dirname(path)
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
    
    if os.path.isfile(file_path):
        try:
            res = read_one(file_path)
            res = json.dumps(res, ensure_ascii=False, indent=4)
            return "读取成功,内容为：" + res
        except Exception as e:
            raise ValueError(f"读取文件失败: {e}")
    elif os.path.isdir(file_path):
        try:
            file_tree = get_file_tree(file_path)
            file_tree = json.dumps(file_tree, ensure_ascii=False, indent=4)
            return file_tree
        except Exception as e:
            raise ValueError(f"获取文件树失败: {e}")
    return "操作成功"


def get_file_tree(folder_path):
    def build_tree(path):
        tree = {}
        for item in os.listdir(path):
            item_path = os.path.join(path, item)
            if os.path.isdir(item_path):
                tree[item] = build_tree(item_path)
            else:
                tree[item] = item
        return tree

    return {folder_path: build_tree(folder_path)}


class files_read_tool:
    @classmethod
    def INPUT_TYPES(s):
        return {
            "required": {
                "folder_path": ("STRING", {"default": "folder path/to/you want to llm control"}),
                "is_enable": ("BOOLEAN", {"default": True}),
            }
        }

    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("tool",)

    FUNCTION = "time"

    # OUTPUT_NODE = False

    CATEGORY = "大模型派对（llm_party）/工具（tools）/实用（Utility）"

    def time(self, folder_path, is_enable=True):
        if is_enable == False:
            return (None,)
        print("Please don't let LLM control your important documents!")
        global global_file_path
        global_file_path=folder_path
        if global_file_path=="" or global_file_path==None:
            print("folder path must be set！")
            return (None,)
        output = [
            {
                "type": "function",
                "function": {
                    "name": "file_read",
                    "description": f"用于读取在{folder_path}下的文件或文件夹。不要读取超出该目录的文件或文件夹！",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "file_path": {
                                "type": "string",
                                "description": "要读取的文件或文件夹，如果是文件夹，则返回该文件夹下的文件树。",
                            }
                        },
                        "required": ["file_path"],
                    },
                },
            }
        ]
        out = json.dumps(output, ensure_ascii=False)
        return (out,)
    @classmethod
    def IS_CHANGED(s):
        # 生成当前时间的哈希值
        hash_value = hashlib.md5(str(datetime.datetime.now()).encode()).hexdigest()
        return hash_value

_TOOL_HOOKS = ["file_read"]
NODE_CLASS_MAPPINGS = {"files_read_tool": files_read_tool}
# 获取系统语言
lang = locale.getlocale()[0]
if 'Chinese' in lang:
   lang = 'zh_CN'
else:
   lang = 'en_US'
import os
import sys
current_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
config_path = os.path.join(current_dir, "config.ini")
import configparser
config = configparser.ConfigParser()
config.read(config_path)
try:
    language = config.get("API_KEYS", "language")
except:
    language = ""
if language == "zh_CN" or language=="en_US":
    lang=language
if lang == "zh_CN":
    NODE_DISPLAY_NAME_MAPPINGS = {"files_read_tool": "本地文件读取工具"}
else:
    NODE_DISPLAY_NAME_MAPPINGS = {"files_read_tool": "local file read tool"}
